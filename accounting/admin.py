from django.contrib import admin
from django.utils.html import format_html, escape
from django.urls import reverse, path
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.utils import timezone
from datetime import timedelta
from .models import (
    Client, ClientPhoneNumber, Supplier, SupplierPhoneNumber,
    Product, Invoice, InvoiceItem, Payment, DashboardStats
)
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.db.models import Sum, Count, Q, F
from django.template.loader import render_to_string
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.contrib import messages
from django.core.cache import cache
from django.shortcuts import redirect
from django.template.response import TemplateResponse
import logging
from django.utils.translation import gettext_lazy as _

logger = logging.getLogger(__name__)

# Inline classes
class PhoneNumberInline(admin.TabularInline):
    extra = 1

class ClientPhoneNumberInline(PhoneNumberInline):
    model = ClientPhoneNumber
    verbose_name = _("Client Phone Number")
    verbose_name_plural = _("Client Phone Numbers")

class SupplierPhoneNumberInline(PhoneNumberInline):
    model = SupplierPhoneNumber
    verbose_name = _("Supplier Phone Number")
    verbose_name_plural = _("Supplier Phone Numbers")

# Admin classes
@admin.register(Client)
class ClientAdmin(admin.ModelAdmin):
    list_display = ('name', 'business_type', 'national_id', 'tax_number', 'num_of_invoices', 'address')
    list_filter = ('business_type',)
    search_fields = ('name', 'national_id', 'tax_number')
    inlines = [ClientPhoneNumberInline]
    fieldsets = (
        (None, {
            'fields': ('name', 'business_type', 'address')
        }),
        (_("Identification"), {
            'fields': ('national_id', 'tax_number'),
            'description': _("National ID for individuals, Tax number for companies")
        }),
        (_("Statistics"), {
            'fields': ('num_of_invoices',),
            'classes': ('collapse',)
        }),
    )
    change_list_template = 'admin/accounting/change_list.html'
    def view_link(self, obj):
        return format_html(
            '<a href="{}" class="button" style="padding: 2px 8px; background: #417690; color: white; border-radius: 3px;">{}</a>',
            reverse('admin:accounting_{}_change'.format(self.model._meta.model_name), args=[obj.id]),
            _("View")
        )
    view_link.short_description = ''

@admin.register(Supplier)
class SupplierAdmin(admin.ModelAdmin):
    list_display = ('name', 'supplier_code', 'business_type', 'tax_number', 'total_orders', 'payment_method')
    list_filter = ('business_type', 'payment_method')
    search_fields = ('name', 'supplier_code', 'tax_number')
    inlines = [SupplierPhoneNumberInline]
    fieldsets = (
        (None, {
            'fields': ('name', 'supplier_code', 'business_type', 'tax_number', 'address')
        }),
        (_("Cooperation"), {
            'fields': ('cooperation_start_date', 'cooperation_end_date', 'website')
        }),
        (_("Payment"), {
            'fields': ('payment_method', 'terms_and_conditions')
        }),
        (_("Statistics"), {
            'fields': ('total_orders',),
            'classes': ('collapse',)
        }),
    )
    change_list_template = 'admin/accounting/change_list.html'
    def view_link(self, obj):
        return format_html(
            '<a href="{}" class="button" style="padding: 2px 8px; background: #417690; color: white; border-radius: 3px;">{}</a>',
            reverse('admin:accounting_{}_change'.format(self.model._meta.model_name), args=[obj.id]),
            _("View")
        )
    view_link.short_description = ''

@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'code', 'wholesale_price', 'retail_price', 'quantity_in_stock', 'reorder_level', 'current_stock_status', 'qr_code_preview', 'supplier')
    list_filter = ('supplier',)
    search_fields = ('name', 'code')
    actions = ['export_to_excel']
    list_editable = ('wholesale_price', 'retail_price', 'reorder_level')
    readonly_fields = ('qr_code_display',)
    fieldsets = (
        (None, {
            'fields': ('name', 'code', 'supplier', 'description')
        }),
        (_("Pricing"), {
            'fields': ('wholesale_price', 'retail_price', 'tax_type', 'tax_value')
        }),
        (_("Inventory"), {
            'fields': ('quantity_in_stock', 'reorder_level')
        }),
        (_("Images"), {
            'fields': ('image', 'qr_code', 'qr_code_display'),
            'description': _("QR code is automatically generated based on product name and code.")
        }),
    )
    change_list_template = 'admin/accounting/change_list.html'
    def qr_code_preview(self, obj):
        if obj.qr_code:
            return format_html(
                '<a href="{}" target="_blank"><img src="{}" width="50" height="50" alt="{}" style="border: 1px solid #ddd; border-radius: 4px;" /></a>',
                escape(obj.qr_code.url), escape(obj.qr_code.url), _("QR Code for") + f" {obj.name}"
            )
        return "-"
    qr_code_preview.short_description = _("QR Code")

    def qr_code_display(self, obj):
        if obj.qr_code:
            return format_html(
                '<a href="{}" target="_blank"><img src="{}" width="100" height="100" alt="{}" style="border: 1px solid #ddd; border-radius: 4px;" /></a>',
                escape(obj.qr_code.url), escape(obj.qr_code.url), _("QR Code for") + f" {obj.name}"
            )
        return _("No QR code available.")
    qr_code_display.short_description = _("QR Code Preview")

    def current_stock_status(self, obj):
        if obj.quantity_in_stock is not None and obj.reorder_level is not None:
            if obj.quantity_in_stock <= obj.reorder_level:
                return format_html('<span style="color: red; font-weight: bold;">{}</span>', _("Low Stock"))
        return format_html('<span style="color: green;">{}</span>', _("In Stock"))
    current_stock_status.short_description = _("Stock Status")

    def export_to_excel(self, request, queryset):
        if not request.user.has_perm('accounting.export_products'):
            self.message_user(request, _("You do not have permission to export products."), level=messages.ERROR)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _("Products")

        headers = [
            _("Name"), _("Code"), _("Wholesale Price"), _("Retail Price"),
            _("Quantity in Stock"), _("Reorder Level"), _("Supplier"), _("Stock Status")
        ]
        ws.append(headers)

        for product in queryset:
            ws.append([
                product.name,
                product.code,
                product.wholesale_price,
                product.retail_price,
                product.quantity_in_stock,
                product.reorder_level,
                str(product.supplier) if product.supplier else "",
                product.current_stock_status
            ])

        bold_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = bold_font
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=products.xlsx'
        wb.save(response)
        return response
    export_to_excel.short_description = _("Export selected products to Excel")
    
    def view_link(self, obj):
        return format_html(
            '<a href="{}" class="button" style="padding: 2px 8px; background: #417690; color: white; border-radius: 3px;">{}</a>',
            reverse('admin:accounting_{}_change'.format(self.model._meta.model_name), args=[obj.id]),
            _("View")
        )
    view_link.short_description = ''

class InvoiceItemInline(admin.TabularInline):
    model = InvoiceItem
    extra = 1
    readonly_fields = ('total',)
    verbose_name = _("Invoice Item")
    verbose_name_plural = _("Invoice Items")

@admin.register(Invoice)
class InvoiceAdmin(admin.ModelAdmin):
    list_display = ('invoice_number', 'invoice_date', 'invoice_type', 'client_or_supplier', 'total_amount', 'payment_status')
    list_filter = ('invoice_type', 'payment_status', 'invoice_date')
    search_fields = ('invoice_number', 'client__name', 'supplier__name')
    inlines = [InvoiceItemInline]
    readonly_fields = ('total_amount',)
    date_hierarchy = 'invoice_date'
    actions = ['export_to_excel', 'mark_as_paid', 'mark_as_unpaid']
    list_per_page = 20
    change_list_template = 'admin/accounting/change_list.html'

    def client_or_supplier(self, obj):
        if obj.invoice_type == 'sale':
            return f"{_('Client')}: {obj.client}"
        return f"{_('Supplier')}: {obj.supplier}"
    client_or_supplier.short_description = _("Client/Supplier")

    def export_to_excel(self, request, queryset):
        if not request.user.has_perm('accounting.export_invoices'):
            self.message_user(request, _("You do not have permission to export invoices."), level=messages.ERROR)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _("Invoices")

        headers = [
            _("Invoice Number"), _("Date"), _("Type"), _("Client/Supplier"),
            _("Total Amount"), _("Payment Status"), _("Items Count")
        ]
        ws.append(headers)

        for invoice in queryset:
            ws.append([
                invoice.invoice_number,
                invoice.invoice_date.strftime('%Y-%m-%d'),
                invoice.get_invoice_type_display(),
                str(invoice.client) if invoice.client else str(invoice.supplier),
                invoice.total_amount,
                invoice.get_payment_status_display(),
                invoice.items.count()
            ])

        bold_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = bold_font
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=invoices.xlsx'
        wb.save(response)
        return response
    export_to_excel.short_description = _("Export selected invoices to Excel")

    def mark_as_paid(self, request, queryset):
        if not request.user.has_perm('accounting.mark_invoices_paid'):
            self.message_user(request, _("You do not have permission to mark invoices as paid."), level=messages.ERROR)
            return
        updated = queryset.update(payment_status='paid')
        self.message_user(request, _("%(count)s invoices marked as paid.") % {'count': updated})
    mark_as_paid.short_description = _("Mark selected invoices as paid")

    def mark_as_unpaid(self, request, queryset):
        if not request.user.has_perm('accounting.mark_invoices_unpaid'):
            self.message_user(request, _("You do not have permission to mark invoices as unpaid."), level=messages.ERROR)
            return
        updated = queryset.update(payment_status='unpaid')
        self.message_user(request, _("%(count)s invoices marked as unpaid.") % {'count': updated})
    mark_as_unpaid.short_description = _("Mark selected invoices as unpaid")
    
    def view_link(self, obj):
        return format_html(
            '<a href="{}" class="button" style="padding: 2px 8px; background: #417690; color: white; border-radius: 3px;">{}</a>',
            reverse('admin:accounting_{}_change'.format(self.model._meta.model_name), args=[obj.id]),
            _("View")
        )
    view_link.short_description = ''

@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = ('invoice', 'amount', 'payment_date', 'payment_method', 'remaining_amount')
    list_filter = ('payment_method', 'payment_date')
    search_fields = ('invoice__invoice_number',)
    readonly_fields = ('remaining_amount',)
    date_hierarchy = 'payment_date'
    change_list_template = 'admin/accounting/change_list.html'

    def remaining_amount(self, obj):
        return obj.remaining_amount
    remaining_amount.short_description = _("Remaining Amount")
    
    def view_link(self, obj):
        return format_html(
            '<a href="{}" class="button" style="padding: 2px 8px; background: #417690; color: white; border-radius: 3px;">{}</a>',
            reverse('admin:accounting_{}_change'.format(self.model._meta.model_name), args=[obj.id]),
            _("View")
        )
    view_link.short_description = ''

@admin.register(DashboardStats)
class DashboardStatsAdmin(admin.ModelAdmin):
    change_list_template = 'admin/dashboard.html'

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def calculate_stats(self, request, from_date=None, to_date=None, client_id=None, supplier_id=None, period='monthly'):
        today = timezone.now().date()
        thirty_days_ago = today - timedelta(days=30)

        sales_queryset = Invoice.objects.filter(invoice_type='sale')
        purchases_queryset = Invoice.objects.filter(invoice_type='purchase')

        # Apply filters
        if from_date:
            try:
                from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
                sales_queryset = sales_queryset.filter(invoice_date__gte=from_date)
                purchases_queryset = purchases_queryset.filter(invoice_date__gte=from_date)
            except ValueError:
                pass
        if to_date:
            try:
                to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
                sales_queryset = sales_queryset.filter(invoice_date__lte=to_date)
                purchases_queryset = purchases_queryset.filter(invoice_date__lte=to_date)
            except ValueError:
                pass
        if client_id:
            sales_queryset = sales_queryset.filter(client_id=client_id)
        if supplier_id:
            purchases_queryset = purchases_queryset.filter(supplier_id=supplier_id)

        total_sales = sales_queryset.aggregate(total=Sum('total_amount'))['total'] or 0
        recent_sales = sales_queryset.filter(invoice_date__gte=thirty_days_ago).aggregate(total=Sum('total_amount'))['total'] or 0
        unpaid_sales = sales_queryset.filter(payment_status='unpaid').aggregate(total=Sum('total_amount'))['total'] or 0

        total_purchases = purchases_queryset.aggregate(total=Sum('total_amount'))['total'] or 0
        recent_purchases = purchases_queryset.filter(invoice_date__gte=thirty_days_ago).aggregate(total=Sum('total_amount'))['total'] or 0
        unpaid_purchases = purchases_queryset.filter(payment_status='unpaid').aggregate(total=Sum('total_amount'))['total'] or 0

        low_stock_products = Product.objects.filter(quantity_in_stock__lte=F('reorder_level'))
        out_of_stock_products = low_stock_products.filter(quantity_in_stock=0).count()

        logger.info(f"Low stock products count: {low_stock_products.count()}")
        for product in low_stock_products:
            logger.debug(f"Product {product.name} (ID: {product.id}, Code: {product.code}): Stock={product.quantity_in_stock}, Reorder={product.reorder_level}")

        low_stock_notifications = [
            {
                'id': product.id,
                'name': product.name,
                'code': product.code,
                'quantity_in_stock': product.quantity_in_stock,
                'reorder_level': product.reorder_level,
                'admin_url': reverse('admin:accounting_product_change', args=[product.id])
            }
            for product in low_stock_products
        ]

        total_clients = Client.objects.count()
        active_clients = Client.objects.filter(invoices__invoice_date__gte=thirty_days_ago).distinct().count()
        total_suppliers = Supplier.objects.count()
        active_suppliers = Supplier.objects.filter(invoices__invoice_date__gte=thirty_days_ago).distinct().count()

        # Fetch recent data with filters applied
        recent_invoices = sales_queryset.union(purchases_queryset).order_by('-id')[:10]
        recent_payments = Payment.objects.filter(
            invoice__in=sales_queryset.values('id').union(purchases_queryset.values('id'))
        ).order_by('-id')[:10]
        recent_orders = sales_queryset.order_by('-id')[:10]  # Use sales invoices as orders

        # Chart data
        if period == 'monthly':
            days = 180
            interval = 'month'
        elif period == 'quarterly':
            days = 365
            interval = 'quarter'
        else:  # yearly
            days = 365 * 3
            interval = 'year'

        months = []
        sales_data = []
        purchases_data = []
        current_date = today - timedelta(days=days)

        while current_date <= today:
            if interval == 'month':
                month_start = current_date.replace(day=1)
                month_end = (month_start + timedelta(days=31)).replace(day=1) - timedelta(days=1)
                label = month_start.strftime('%b %Y')
            elif interval == 'quarter':
                quarter = (current_date.month - 1) // 3 + 1
                month_start = current_date.replace(month=((quarter - 1) * 3 + 1), day=1)
                month_end = (month_start + timedelta(days=90)).replace(day=1) - timedelta(days=1)
                label = f"Q{quarter} {month_start.year}"
            else:  # yearly
                month_start = current_date.replace(month=1, day=1)
                month_end = current_date.replace(month=12, day=31)
                label = month_start.strftime('%Y')

            month_sales = sales_queryset.filter(
                invoice_date__gte=month_start,
                invoice_date__lte=month_end
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            month_purchases = purchases_queryset.filter(
                invoice_date__gte=month_start,
                invoice_date__lte=month_end
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            months.append(label)
            sales_data.append(float(month_sales))
            purchases_data.append(float(month_purchases))

            if interval == 'month':
                current_date = month_end + timedelta(days=1)
            elif interval == 'quarter':
                current_date = month_end + timedelta(days=1)
            else:
                current_date = current_date.replace(year=current_date.year + 1)

        chart_data = {
            'type': 'line',
            'data': {
                'labels': months,
                'datasets': [
                    {
                        'label': _('Sales'),
                        'data': sales_data,
                        'borderColor': '#007bff',
                        'backgroundColor': 'rgba(0, 123, 255, 0.1)',
                        'fill': True
                    },
                    {
                        'label': _('Purchases'),
                        'data': purchases_data,
                        'borderColor': '#28a745',
                        'backgroundColor': 'rgba(40, 167, 69, 0.1)',
                        'fill': True
                    }
                ]
            },
            'options': {
                'responsive': True,
                'scales': {
                    'y': {
                        'beginAtZero': True,
                        'title': {'display': True, 'text': _('Amount ($)')}
                    },
                    'x': {
                        'title': {'display': True, 'text': _(interval.capitalize())}
                    }
                }
            }
        }

        return {
            'total_sales': total_sales,
            'recent_sales': recent_sales,
            'unpaid_sales': unpaid_sales,
            'total_purchases': total_purchases,
            'recent_purchases': recent_purchases,
            'unpaid_purchases': unpaid_purchases,
            'low_stock_products': low_stock_products.count(),
            'out_of_stock_products': out_of_stock_products,
            'total_clients': total_clients,
            'active_clients': active_clients,
            'total_suppliers': total_suppliers,
            'active_suppliers': active_suppliers,
            'recent_invoices': recent_invoices,
            'recent_payments': recent_payments,
            'recent_orders': recent_orders,
            'chart_data': chart_data,
            'low_stock_notifications': low_stock_notifications,
        }

    def changelist_view(self, request, extra_context=None):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        client_id = request.GET.get('client')
        supplier_id = request.GET.get('supplier')
        period = request.GET.get('period', 'monthly')

        cache_key = f'dashboard_stats_{from_date}_{to_date}_{client_id}_{supplier_id}_{period}'
        stats = cache.get(cache_key)
        if not stats:
            stats = self.calculate_stats(request, from_date, to_date, client_id, supplier_id, period)
            cache.set(cache_key, stats, timeout=300)

        extra_context = extra_context or {}
        extra_context.update(stats)
        extra_context.update({
            'sales_report_url': reverse('admin:accounting_sales_report'),
            'purchases_report_url': reverse('admin:accounting_purchases_report'),
            'clients': Client.objects.all(),
            'suppliers': Supplier.objects.all(),
            'from_date': from_date or '',
            'to_date': to_date or '',
            'selected_client': client_id or '',
            'selected_supplier': supplier_id or '',
            'app_label': 'accounting',
        })

        return TemplateResponse(
            request,
            self.change_list_template,
            extra_context
        )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('dashboardstats/', self.admin_site.admin_view(self.changelist_view), name='accounting_dashboardstats'),
            path('sales-report/', self.admin_site.admin_view(self.sales_report), name='accounting_sales_report'),
            path('purchases-report/', self.admin_site.admin_view(self.purchases_report), name='accounting_purchases_report'),
            path('chart-data/', self.admin_site.admin_view(self.chart_data), name='accounting_chart_data'),
        ]
        return custom_urls + urls

    def chart_data(self, request):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        client_id = request.GET.get('client')
        supplier_id = request.GET.get('supplier')
        period = request.GET.get('period', 'monthly')

        stats = self.calculate_stats(request, from_date, to_date, client_id, supplier_id, period)
        return JsonResponse({'chart_data': stats['chart_data']})

    def sales_report(self, request):
        if not request.user.has_perm('accounting.view_sales_report'):
            self.message_user(request, _("You do not have permission to view sales reports."), level=messages.ERROR)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        client_id = request.GET.get('client')

        queryset = Invoice.objects.filter(invoice_type='sale')
        if from_date:
            queryset = queryset.filter(invoice_date__gte=from_date)
        if to_date:
            queryset = queryset.filter(invoice_date__lte=to_date)
        if client_id:
            queryset = queryset.filter(client_id=client_id)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{_("sales_report")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            _("Invoice Number"), _("Date"), _("Client"), _("Total Amount"),
            _("Payment Status"), _("Items Count")
        ])

        for invoice in queryset:
            writer.writerow([
                invoice.invoice_number,
                invoice.invoice_date.strftime('%Y-%m-%d'),
                str(invoice.client) if invoice.client else "",
                invoice.total_amount,
                invoice.get_payment_status_display(),
                invoice.items.count()
            ])

        return response

    def purchases_report(self, request):
        if not request.user.has_perm('accounting.view_purchases_report'):
            self.message_user(request, _("You do not have permission to view purchases reports."), level=messages.ERROR)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        supplier_id = request.GET.get('supplier')

        queryset = Invoice.objects.filter(invoice_type='purchase')
        if from_date:
            queryset = queryset.filter(invoice_date__gte=from_date)
        if to_date:
            queryset = queryset.filter(invoice_date__lte=to_date)
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{_("purchases_report")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            _("Invoice Number"), _("Date"), _("Supplier"), _("Total Amount"),
            _("Payment Status"), _("Items Count")
        ])

        for invoice in queryset:
            writer.writerow([
                invoice.invoice_number,
                invoice.invoice_date.strftime('%Y-%m-%d'),
                str(invoice.supplier) if invoice.supplier else "",
                invoice.total_amount,
                invoice.get_payment_status_display(),
                invoice.items.count()
            ])

        return response

# Custom Admin Site
class CustomAdminSite(admin.AdminSite):
    site_header = _("Accounting System Dashboard")
    site_title = _("Accounting Dashboard")

    def index(self, request, extra_context=None):
        dashboard_admin = DashboardStatsAdmin(DashboardStats, self)
        extra_context = extra_context or {}
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        client_id = request.GET.get('client')
        supplier_id = request.GET.get('supplier')
        period = request.GET.get('period', 'monthly')

        cache_key = f'admin_dashboard_stats_{from_date}_{to_date}_{client_id}_{supplier_id}_{period}'
        stats = cache.get(cache_key)
        if not stats:
            stats = dashboard_admin.calculate_stats(request, from_date, to_date, client_id, supplier_id, period)
            cache.set(cache_key, stats, timeout=300)

        extra_context.update(stats)
        extra_context.update({
            'clients': Client.objects.all(),
            'suppliers': Supplier.objects.all(),
            'from_date': from_date or '',
            'to_date': to_date or '',
            'selected_client': client_id or '',
            'selected_supplier': supplier_id or '',
            'app_label': 'accounting'
        })
        return TemplateResponse(
            request,
            'admin/dashboard.html',
            extra_context
        )

custom_admin_site = CustomAdminSite(name='custom_admin')
custom_admin_site.register(Client, ClientAdmin)
custom_admin_site.register(Supplier, SupplierAdmin)
custom_admin_site.register(Product, ProductAdmin)
custom_admin_site.register(Invoice, InvoiceAdmin)
custom_admin_site.register(Payment, PaymentAdmin)
custom_admin_site.register(DashboardStats, DashboardStatsAdmin)